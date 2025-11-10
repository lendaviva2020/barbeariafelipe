"""
Admin Reports - PDF/Excel Export
Extraído da estrutura do React Reports.tsx
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from agendamentos.models import Agendamento
from servicos.models import Servico
from barbeiros.models import Barbeiro

class ReportsRevenueView(APIView):
    """Relatório de faturamento"""
    permission_classes = (IsAdminUser,)
    
    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        agendamentos = Agendamento.objects.filter(
            appointment_date__range=[start_date, end_date],
            status='completed'
        )
        
        revenue_by_day = {}
        for apt in agendamentos:
            date_key = str(apt.appointment_date)
            if date_key not in revenue_by_day:
                revenue_by_day[date_key] = 0
            revenue_by_day[date_key] += float(apt.price)
        
        return Response({
            'total_revenue': agendamentos.aggregate(Sum('price'))['price__sum'] or 0,
            'by_day': revenue_by_day,
            'payment_methods': agendamentos.values('payment_method').annotate(
                count=Count('id'),
                total=Sum('price')
            )
        })

class ReportsServicesView(APIView):
    """Relatório de serviços"""
    permission_classes = (IsAdminUser,)
    
    def get(self, request):
        services = Servico.objects.all()
        
        service_stats = []
        for service in services:
            appointments = Agendamento.objects.filter(service=service)
            service_stats.append({
                'id': service.id,
                'name': service.name,
                'total_appointments': appointments.count(),
                'completed': appointments.filter(status='completed').count(),
                'revenue': appointments.filter(status='completed').aggregate(Sum('price'))['price__sum'] or 0
            })
        
        return Response({
            'services': sorted(service_stats, key=lambda x: x['total_appointments'], reverse=True)
        })

class ReportsBarbersPerformanceView(APIView):
    """Relatório de performance dos barbeiros"""
    permission_classes = (IsAdminUser,)
    
    def get(self, request):
        barbers = Barbeiro.objects.filter(active=True)
        
        barber_stats = []
        for barber in barbers:
            appointments = Agendamento.objects.filter(barber=barber)
            completed = appointments.filter(status='completed')
            
            barber_stats.append({
                'id': barber.id,
                'name': barber.name,
                'total_appointments': appointments.count(),
                'completed': completed.count(),
                'revenue': completed.aggregate(Sum('price'))['price__sum'] or 0,
                'avg_rating': 4.8  # Mock - seria calculado de reviews
            })
        
        return Response({
            'barbers': sorted(barber_stats, key=lambda x: x['revenue'], reverse=True)
        })

class ExportPDFView(APIView):
    """Exportar relatório em PDF com reportlab"""
    permission_classes = (IsAdminUser,)
    
    def post(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.units import cm
        from io import BytesIO
        import tempfile
        
        # Parâmetros
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        report_type = request.data.get('type', 'revenue')
        
        # Criar buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title='Relatório Barbearia Francisco')
        
        # Elementos do PDF
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#8B2635'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        elements.append(Paragraph('Barbearia Francisco', title_style))
        elements.append(Paragraph(f'Relatório de {report_type.title()}', styles['Heading2']))
        elements.append(Paragraph(f'Período: {start_date} a {end_date}', styles['Normal']))
        elements.append(Spacer(1, 1*cm))
        
        # Buscar dados
        agendamentos = Agendamento.objects.filter(
            appointment_date__range=[start_date, end_date],
            status='completed'
        )
        
        total_revenue = agendamentos.aggregate(Sum('price'))['price__sum'] or 0
        
        # Tabela de resumo
        data = [
            ['Métrica', 'Valor'],
            ['Total de Agendamentos', str(agendamentos.count())],
            ['Faturamento Total', f'R$ {total_revenue:.2f}'],
            ['Ticket Médio', f'R$ {(total_revenue/agendamentos.count() if agendamentos.count() > 0 else 0):.2f}'],
        ]
        
        table = Table(data, colWidths=[8*cm, 6*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C9A961')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Footer
        elements.append(Paragraph(
            f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}',
            styles['Normal']
        ))
        
        # Gerar PDF
        doc.build(elements)
        
        # Retornar response
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{report_type}_{start_date}_{end_date}.pdf"'
        
        return response

class ExportExcelView(APIView):
    """Exportar relatório em Excel com openpyxl"""
    permission_classes = (IsAdminUser,)
    
    def post(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Parâmetros
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        
        # Criar workbook
        wb = Workbook()
        
        # ========== ABA 1: FATURAMENTO ==========
        ws_revenue = wb.active
        ws_revenue.title = "Faturamento"
        
        # Headers
        headers = ['Data', 'Cliente', 'Serviço', 'Barbeiro', 'Valor', 'Pagamento']
        ws_revenue.append(headers)
        
        # Estilizar header
        header_fill = PatternFill(start_color='C9A961', end_color='C9A961', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col in range(1, len(headers) + 1):
            cell = ws_revenue.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        agendamentos = Agendamento.objects.filter(
            appointment_date__range=[start_date, end_date],
            status='completed'
        ).select_related('service', 'barber')
        
        for apt in agendamentos:
            ws_revenue.append([
                str(apt.appointment_date),
                apt.customer_name,
                apt.service.name if apt.service else '',
                apt.barber.name if apt.barber else '',
                float(apt.price),
                apt.payment_method
            ])
        
        # Totais
        row_total = ws_revenue.max_row + 2
        ws_revenue.cell(row_total, 4).value = 'TOTAL:'
        ws_revenue.cell(row_total, 4).font = Font(bold=True)
        ws_revenue.cell(row_total, 5).value = f'=SUM(E2:E{ws_revenue.max_row-1})'
        ws_revenue.cell(row_total, 5).font = Font(bold=True)
        ws_revenue.cell(row_total, 5).number_format = 'R$ #,##0.00'
        
        # Auto-width
        for col in range(1, len(headers) + 1):
            ws_revenue.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== ABA 2: SERVIÇOS ==========
        ws_services = wb.create_sheet("Serviços")
        
        services_data = Servico.objects.all()
        ws_services.append(['Serviço', 'Categoria', 'Preço', 'Duração (min)', 'Total Agendamentos'])
        
        # Header style
        for col in range(1, 6):
            cell = ws_services.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
        
        for service in services_data:
            count = Agendamento.objects.filter(service=service).count()
            ws_services.append([
                service.name,
                service.category or '',
                float(service.price),
                service.duration,
                count
            ])
        
        # ========== ABA 3: BARBEIROS ==========
        ws_barbers = wb.create_sheet("Barbeiros")
        
        barbers_data = Barbeiro.objects.filter(active=True)
        ws_barbers.append(['Barbeiro', 'Email', 'Telefone', 'Agendamentos', 'Faturamento'])
        
        # Header style
        for col in range(1, 6):
            cell = ws_barbers.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
        
        for barber in barbers_data:
            barber_apts = Agendamento.objects.filter(barber=barber, status='completed')
            revenue = barber_apts.aggregate(Sum('price'))['price__sum'] or 0
            
            ws_barbers.append([
                barber.name,
                barber.email or '',
                barber.phone or '',
                barber_apts.count(),
                float(revenue)
            ])
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Retornar response
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_completo_{start_date}_{end_date}.xlsx"'
        
        return response

